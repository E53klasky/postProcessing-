"""
streamline_radial_dist.py
Reads three ADIOS2 BP files (ground-truth, low-res, compressed).
coords_x and coords_y are arrays of length 2500 — one (x,y) pair per RK step.

Computes two quantities from the seed at each RK step:
  - Radial distance  r     = sqrt((x-sx)^2 + (y-sy)^2)
  - Signed angle     theta = signed angle (degrees) between [1,0] and [x-sx, y-sy]
                             positive = above seed (y > sy), negative = below

Output: one Excel file with TWO sheets, each 2501 rows (header + 2500):

  Sheet 1 "Radial Distances":
      RK Step | dist_ground_truth | dist_low_res | dist_compressed

  Sheet 2 "Theta (degrees)":
      RK Step | theta_ground_truth | theta_low_res | theta_compressed

Usage:
    python streamline_radial_dist.py \
        --gt gt.bp --low_res lowres.bp --compressed compressed.bp
"""

import os
import argparse
import numpy as np
from rich.traceback import install
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ReaderClass import Reader


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_arguments():
    p = argparse.ArgumentParser()
    p.add_argument("--gt",            "-gt", type=str, required=True)
    p.add_argument("--low_res",       "-lr", type=str, required=True)
    p.add_argument("--compressed",    "-c",  type=str, required=True)
    p.add_argument("--gt_io",         type=str, default="reader_gt")
    p.add_argument("--low_res_io",    type=str, default="reader_lr")
    p.add_argument("--compressed_io", type=str, default="reader_compressed")
    p.add_argument("--xml",  "-x",    type=str, default=None)
    p.add_argument("--var_x",         type=str, default="coords_x")
    p.add_argument("--var_y",         type=str, default="coords_y")
    p.add_argument("--seed_x",        type=float, default=0.2)
    p.add_argument("--seed_y",        type=float, default=0.2)
    p.add_argument("--output_dir",    type=str, default="../RESULTS")
    p.add_argument("--output_file",   type=str, default="streamline_distances.xlsx")
    return p.parse_args()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def read_array(reader, var):
    reader.set_read_vars([var])
    return reader.read_step(var)


def radial_distance(x, y, sx, sy):
    """Euclidean distance from seed at each RK step."""
    return np.sqrt((x - sx)**2 + (y - sy)**2)


def signed_theta_degrees(x, y, sx, sy):
    """
    Signed angle in degrees between the reference vector [1, 0] and the
    displacement vector [x-sx, y-sy], measured at each RK step.

    Method:
        a = [1, 0]   (fixed reference along +x from seed)
        b = [x-sx, y-sy]

        angle = arccos( a·b / (|a| |b|) )   <- always in [0, 180]

        Sign: positive if y > sy  (point is above the seed's x-axis)
              negative if y < sy  (point is below)
              zero     if y == sy
    """
    dx = x - sx
    dy = y - sy

    # |b| — protect against zero-length vector (seed point itself)
    mag_b = np.sqrt(dx**2 + dy**2)
    mag_b = np.where(mag_b == 0, np.nan, mag_b)

    # a·b = 1*dx + 0*dy = dx,  |a| = 1
    cos_theta = dx / mag_b                          # clamp for floating-point safety
    cos_theta = np.clip(cos_theta, -1.0, 1.0)

    angle_deg = np.degrees(np.arccos(cos_theta))   # [0, 180]

    # Apply sign: negative below the seed's horizontal
    angle_deg = np.where(dy < 0, -angle_deg, angle_deg)

    return angle_deg


def write_sheet(ws, header, rows):
    """Write header + rows to an openpyxl worksheet."""
    ws.append(header)
    for row in rows:
        ws.append(row)
    # Auto-width (generous defaults)
    col_widths = [10] + [20] * (len(header) - 1)
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_arguments()
    os.makedirs(args.output_dir, exist_ok=True)

    # ---- Open readers ----
    reader_gt = Reader(args.gt_io,         args.gt,         xml=args.xml)
    reader_lr = Reader(args.low_res_io,    args.low_res,    xml=args.xml)
    reader_c  = Reader(args.compressed_io, args.compressed, xml=args.xml)

    reader_gt.begin_step()
    reader_lr.begin_step()
    reader_c.begin_step()

    x_gt = read_array(reader_gt, args.var_x)
    y_gt = read_array(reader_gt, args.var_y)
    x_lr = read_array(reader_lr, args.var_x)
    y_lr = read_array(reader_lr, args.var_y)
    x_c  = read_array(reader_c,  args.var_x)
    y_c  = read_array(reader_c,  args.var_y)

    reader_gt.end_step(); reader_gt.close()
    reader_lr.end_step(); reader_lr.close()
    reader_c.end_step();  reader_c.close()

    sx, sy = args.seed_x, args.seed_y

    # ---- Radial distances ----
    d_gt = radial_distance(x_gt, y_gt, sx, sy)
    d_lr = radial_distance(x_lr, y_lr, sx, sy)
    d_c  = radial_distance(x_c,  y_c,  sx, sy)

    # ---- Signed theta (degrees) ----
    t_gt = signed_theta_degrees(x_gt, y_gt, sx, sy)
    t_lr = signed_theta_degrees(x_lr, y_lr, sx, sy)
    t_c  = signed_theta_degrees(x_c,  y_c,  sx, sy)

    n = len(d_gt)
    print(f"RK steps found: {n}")

    # ---- Build Excel workbook — single sheet ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Streamline Data"

    header = [
        "RK Step",
        "dist_ground_truth", "dist_low_res", "dist_compressed",
        "theta_ground_truth", "theta_low_res", "theta_compressed",
    ]
    rows = [
        [i,
         float(d_gt[i]), float(d_lr[i]), float(d_c[i]),
         float(t_gt[i]), float(t_lr[i]), float(t_c[i])]
        for i in range(n)
    ]
    write_sheet(ws, header, rows)

    out_path = os.path.join(args.output_dir, args.output_file)
    wb.save(out_path)
    print(f"Saved -> {out_path}")


if __name__ == "__main__":
    install()
    main()
